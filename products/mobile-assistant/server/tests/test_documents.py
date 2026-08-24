from pathlib import Path
import re
from zipfile import ZIP_DEFLATED, ZipFile

from miru_server.documents import extract


def test_extract_xlsx_with_sheet_values(tmp_path: Path):
    import openpyxl

    path = tmp_path / "sales.xlsx"
    book = openpyxl.Workbook()
    book.active.title = "销售"
    book.active.append(["月份", "营收"])
    book.active.append(["八月", 120])
    book.save(path)

    result = extract(path, "spreadsheet")
    assert "工作表：销售" in result.text
    assert "八月 | 120" in result.text
    assert not result.error


def test_extract_xlsx_with_empty_dimensionless_cover_and_long_detail(tmp_path: Path):
    """兼容 Cashew：空白 Sheet1 无 dimension，真实数据在后续工作表且超过 80 行。"""
    import openpyxl

    path = tmp_path / "cashew-style.xlsx"
    book = openpyxl.Workbook()
    book.active.title = "Sheet1"
    details = book.create_sheet("Transaction Details")
    details.append(["账户", "数量", "标题"])
    for index in range(1, 121):
        details.append(["现金", -index, f"第 {index} 笔"])
    summary = book.create_sheet("Monthly Summary")
    summary.append(["Month", "Income", "Expense", "Net"])
    summary.append(["2026-08", 2500, 3251, -751])
    book.save(path)

    # 模拟部分第三方导出器省略空白工作表的 dimension 标签。
    rewritten = tmp_path / "rewritten.xlsx"
    with ZipFile(path, "r") as source, ZipFile(rewritten, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(br"<dimension[^>]*/>", b"", data)
            target.writestr(item, data)
    rewritten.replace(path)

    result = extract(path, "spreadsheet")
    assert not result.error
    assert "工作表：Monthly Summary" in result.text
    assert "2026-08 | 2500 | 3251 | -751" in result.text
    assert "第 120 笔" in result.text


def test_extract_pdf_creates_page_preview(tmp_path: Path):
    import fitz

    path = tmp_path / "note.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "Miru PDF test")
    document.save(path)
    document.close()

    result = extract(path, "document", max_preview_pages=1)
    assert "Miru PDF test" in result.text
    assert len(result.previews) == 1
    assert Path(result.previews[0]).exists()
